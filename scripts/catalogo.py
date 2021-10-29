from .core.web import Web
import sys
from os.path import isfile
import openpyxl
from openpyxl.utils import range_boundaries
from .core.util import write, clean_url, read
from textwrap import dedent
import re
from munch import Munch
from io import BytesIO
import unidecode


web = Web()
re_sp = re.compile(r"\s+")
re_dash = re.compile(r"\s+-|-\s+")

def get_ods(url, ext):
    web.get(url)
    for url in web.soup.select("article.mod-detail li a"):
        url = url.attrs.get("href")
        if url is not None and url.endswith(ext):
            return url

class Item(Munch):
    def __init__(self, *args, **kargv):
        super().__init__(*args, **kargv)
        for a, b in (
            (r"Administración General del Estado|A\.G\.E\.", "AGE"),
            (r"Administraciones públicas|AA\.PP\.", "AAPP"),
            (r"Esquema Nacional de Interoperabilidad( \(ENI\))?", "ENI"),
            ("Esquema Nacional de Seguridad( \(ENS\))?", "ENS"),
        ):
            self.descripcion = re.sub(a, b, self.descripcion, flags=re.IGNORECASE)

class Catalogo:
    def __init__(self, url="https://www.administracionelectronica.gob.es/pae_Home/pae_Estrategias/Racionaliza_y_Comparte/catalogo-servicios-admon-digital.html"):
        self.root = url
        self.version={}
        for ext in ("xlsx", "pdf", "ods"):
            url = get_ods(self.root, "."+ext)
            if url:
                self.version[ext.upper()] = url
        if self.version.get("XLSX") is None:
            raise Exception("xlsx no encontrado en "+self.root)
        print(self.version['XLSX'])
        self.content = BytesIO(web.s.get(self.version['XLSX']).content)
        self.book = openpyxl.load_workbook(filename=self.content)

    @property
    def sheets(self):
        arr = []
        for number, sheet in enumerate(self.book):
            for cell_group in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
                vals = set()
                for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row, values_only=True):
                    for cell in row:
                        if cell is not None:
                            vals.add(cell.strip())
                sheet.unmerge_cells(str(cell_group))
                if vals:
                    top_left_cell_value = vals.pop()
                    for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                        for cell in row:
                            cell.value = top_left_cell_value
            cap = []
            head = True
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                index = index + 1
                text = " ".join(c for c in row if c is not None)
                text = re_sp.sub(" ", text).strip().lower()
                if len(text)==0:
                    continue
                c1 = (row[0] or "")
                c1 = c1.strip()
                s1 = re_sp.sub(" ", c1).strip().lower()
                if head and len(s1)==0:
                    continue
                if head:
                    if s1 in ("categoría", "información básica", "tipo de solución"):
                        if len(cap)==0:
                            continue
                        head = False
                        arr.append(Munch(
                            number=number,
                            title=cap[0],
                            description="\n".join(cap[1:]),
                            data=sheet,
                            start=index,
                            end=index+1,
                            min_col=1
                        ))
                        cap = None
                    else:
                        cap.extend(s.strip() for s in c1.split("\n") if s.strip())
                if arr:
                    c2 = (row[1] or "")
                    c2 = c2.strip()
                    s2 = re_sp.sub(" ", c2).strip().lower()
                    if s2 == "tipo de solución":
                        arr[-1].min_col=2
                    if s1 == "tipo de solución":
                        arr[-1].start=index+1
                    if not head:
                        arr[-1].end=index

        def myindex(sheet, arr, field, *args):
            if field in sheet:
                return
            for e in args:
                e = e.lower()
                if e in arr:
                    sheet[field] = arr.index(e)
                    return

        for sheet in arr:
            for row in sheet.data.iter_rows(min_col=sheet.min_col, min_row=sheet.start, max_row=sheet.end, values_only=True):
                if row[0] is not None:
                    break
                sheet.start = sheet.start + 1
            for row in sheet.data.iter_rows(min_col=sheet.min_col, max_row=sheet.start, values_only=True):
                row = [(r or "").strip().lower() for r in row]
                myindex(sheet, row, "ambito", "ámbito")
                myindex(sheet, row, "organismo", "organismo")
                myindex(sheet, row, "enlace", "enlace + información")
                myindex(sheet, row, "comun", "Servicio declarado como compartido (AGE)", "Servicio declarado como compartido")
            print(sheet.data.title, sheet.start, sheet.end)
            sheet.rows = []
            tipo = None
            for row in sheet.data.iter_rows(min_col=sheet.min_col, min_row=sheet.start, max_row=sheet.end, values_only=True):
                row = [(r or "").strip() for r in row]
                if row[0] not in (None, ""):
                    tipo = row[0]
                nombre, desc = row[1:3]
                nombre = nombre.rstrip(".")
                nombre = re_dash.sub(" - ", nombre)
                nombre = nombre.replace("  ", " ")
                urls = [u for u in re.split(r"\s+", row[sheet.enlace]) if u.startswith("http")]
                ambito = row[sheet.ambito]
                ambito = re.sub(r"[Tt]odas (las )?AAPP", "AAPP", ambito)
                organismo = row[sheet.organismo]
                if desc and not desc.endswith("."):
                    desc = desc + "."
                comun = False
                if sheet.get("comun") is not None:
                    comun = row[sheet.comun].lower() in ("si", "sí")
                sheet.rows.append(Item(
                    tipo = tipo,
                    nombre = nombre,
                    ambito = ambito,
                    organismo = organismo,
                    urls=urls,
                    descripcion=desc,
                    comun=comun
                ))
            #sheet.rows=list(range(sheet.start+1, sheet.end+1))
            #sheet.data.filter(row_indices=sheet.rows)
            #sheet.data.name_columns_by_row(sheet.start)
        return arr

    def save(self, salida):
        version = map(lambda x: '"{}":"{}"'.format(*x), self.version.items())
        version = ", ".join(version)
        MD = [dedent('''
        ---
        title: Catálogo de servicios comunes
        summary: "Fuente: [PAe - Catálogo de servicios]({root})"
        version: {{ {version} }}
        replace:
          "☑": '<abbr class="nou comun" title="Servicio declarado como compartido">☑</abbr>'
          "☐": '<abbr class="nou nocom" title="Servicio NO declarado como compartido">☐</abbr>'
        ---
        ''').format(root=self.root, version=version).strip()+"\n"]
        MD.append(dedent('''
        <script>
            function getBloque(selector) {
                var bloque = [];
                document.querySelectorAll(selector).forEach(function(e) {
                    let p=e.parentNode;
                    bloque.push(p);
                    while (p.nextElementSibling && p.nextElementSibling.tagName=="P" && p.nextElementSibling.querySelector("abbr.nou")==null) {
                        p = p.nextElementSibling;
                        bloque.push(p);
                    }
                });
                return bloque;
            }
            function onlyComun() {
                const hide = document.getElementById("comun").checked;
                getBloque("abbr.nocom").forEach(function(e) {
                    if (hide) e.setAttribute("style", "display: none;");
                    else e.removeAttribute("style");
                });
            }
        </script>

        <input type="checkbox" id="comun" onclick="onlyComun(this)"/> <label for="comun">Ver solo servicios declarados como compartidos</label>
        ''').lstrip())
        for sheet in self.sheets:
            MD.append("# "+sheet.title+"\n")
            if sheet.description:
                MD.append(sheet.description+"\n")
            last_tipo = None
            for row in sheet.rows:
                if row.tipo != last_tipo:
                    MD.append("## "+row.tipo+"\n")
                last_tipo = row.tipo
                if len(row.urls)>0:
                    row.nombre="[{}]({})".format(row.nombre, row.urls[0])
                if len(row.urls)>1:
                    if not row.descripcion.endswith("."):
                        row.descripcion = row.descripcion+"."
                    urls=map(lambda x:"[{}]({})".format(clean_url(x), x), row.urls[1:])
                    row.descripcion = row.descripcion + " " + ", ".join(urls)
                if row.comun:
                    check = "<abbr class='nou comun' title='Servicio declarado como compartido'>&#9745;</abbr> "
                    check = "☑ "
                else:
                    check = "<abbr class='nou nocom' title='Servicio NO declarado como compartido'>&#9744;</abbr> "
                    check = "☐ "
                MD.append(check+"**{nombre}**: {descripcion}\n".format(**row))
                row.organismo = ", ".join(o.strip() for o in row.organismo.split("\n"))
                MD.append("Ambito: {ambito} || Organismo: {organismo}\n".format(**row))

        write(salida, "\n".join(MD))

if __name__ == "__main__":
    c = Catalogo()
    c.save("content/posts/otros/catalogo.md")
